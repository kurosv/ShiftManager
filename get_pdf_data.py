# TODO
#   ・シフトデータExcel書き込み
#   ・シフトデータから当月勤務地一覧List->書き込み
#   ・日付書き込み
#   ・Excelファイル名変更
#   ・Excelシート操作（名前変更、前月シートコピーなど）
#   ・外部カレンダー連携・登録（Google or Yahoo）
#   ・シフトデータDB登録

# TODO 優先度↓
#   ・祝日一覧更新

# TODO 最終的には Py管理->最終提出用Excel出力 にしたい

def_dir = 'G:\\マイドライブ\\ONESTOP'
excel_file = '勤怠報告書_須貝_'


# ----- inner proccess -----
def convert_shift(locate):
    if locate == '':
        return ''

    shift_ptn = {
        '登別': '登別 三愛病院',
        '旭町': '旭町調剤薬局',
        '川沿': '川沿調剤薬局',
        '若草': '若草調剤薬局',
        '若草PM': '若草調剤薬局',
        'アルビ': 'アルビ 苫小牧',
        'アルビ岩': 'アルビ 岩見沢',
        'クロ新川': 'クローバー薬局 新川店',
        '百合が原': 'クローバー薬局 百合が原店',
        '岩内': '菜の花調剤薬局(岩内)',
    }
    return shift_ptn[locate]
# --------------------------


def read_file():
    from tkinter import filedialog

    typ = [('PDF', '*.pdf')]
    fp = filedialog.askopenfilename(filetypes=typ, initialdir=def_dir)

    return fp


def read_pdf_data(fp):
    import tabula
    import pandas as pd

    dfs = tabula.read_pdf(fp,
                          lattice=True,
                          pandas_options={'header': 0})

    df = dfs[0]
    df.fillna('', inplace=True)
    df.set_index('日付', inplace=True)
    return df


def write_excel(df):
    import openpyxl
    # シフトデータ指定
    col_idx = "須貝"
    # 書き込み行
    row_start = 2
    # 書き込み列
    col_date = 1
    col_shift = 3
    col_wl = 13


    # df列抽出
    my_shift = df[col_idx]
    shift_date = df.index.values
    work_location = list(set(my_shift))
    work_location.remove('')

    # Excel読み込み
    wb = openpyxl.load_workbook('C:\\Users\\user\\Desktop\\shift_test.xlsx')
    ws = wb['22.11']

    # 日付書き込み
    import re
    d = re.sub(r'([0-9]+)月([0-9]+)日', r'\1/\2', shift_date[0])
    ws.cell(row_start, col_date, value=d)

    # 勤務地書き込み
    for i in range(len(work_location)):
        ws.cell(row_start + i, col_wl, value=work_location[i])

    # シフト書き込み
    for i in range(len(my_shift)):
        ws.cell(row_start+i, col_shift, value=convert_shift(my_shift[i]))


    # 別名で保存
    wb.save('C:\\Users\\user\\Desktop\\shift_test.xlsx')


    return




def main():
    # ファイル読み込み
    # PDF読み込み
    df = read_pdf_data(read_file())
    # Excel書き込み
    write_excel(df)


if __name__ == '__main__':
    main()
